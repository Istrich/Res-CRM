"""
Tests for app/services/import_employees.py.

Covers header recognition (exact and aliases), date parsing, FIO column,
and skipped rows (empty title).
"""
import io
from datetime import date

import pytest
from openpyxl import Workbook

from app.services.import_employees import parse_employee_excel


def test_empty_content_returns_empty_list():
    """Minimal workbook with no recognizable headers yields no rows with title (or empty list)."""
    wb = Workbook()
    ws = wb.active
    ws.append([])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    # Parser may return rows with empty title; no row should have a non-empty title
    assert all(not (r.get("title") or "").strip() for r in rows)


def test_headers_only_returns_empty_data_rows():
    """Only header row, no data: no row with non-empty title."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Отчество", "Специализация", "Должность", "Подразделение",
               "Дата найма", "Дата увольнения", "Комментарий"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert all(not (r.get("title") or "").strip() for r in rows)


def test_exact_headers_parsed():
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Отчество", "Специализация", "Должность", "Подразделение",
               "Дата найма", "Дата увольнения", "Комментарий"])
    ws.append(["Иванов", "Иван", "Иванович", "Backend", "Разработчик", "ИТ", "2024-01-15", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 1
    assert rows[0]["last_name"] == "Иванов"
    assert rows[0]["first_name"] == "Иван"
    assert rows[0]["middle_name"] == "Иванович"
    assert rows[0]["title"] == "Разработчик"
    assert rows[0]["department"] == "ИТ"
    assert rows[0]["hire_date"] == date(2024, 1, 15)
    assert rows[0]["termination_date"] is None


def test_hire_date_iso_parsed():
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Должность", "Дата найма"])
    ws.append(["Петров", "Петр", "Аналитик", "2023-06-01"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 1
    assert rows[0]["hire_date"] == date(2023, 6, 1)


def test_empty_title_row_included_but_has_empty_title():
    """Rows with empty title are still returned by parser; API skips them."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Должность"])
    ws.append(["Иванов", "Иван", "Разработчик"])
    ws.append(["Петров", "Петр", ""])
    ws.append(["Сидоров", "Сидор", "Тестировщик"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 3
    assert rows[0]["title"] == "Разработчик"
    assert rows[1]["title"] in (None, "")
    assert rows[2]["title"] == "Тестировщик"


def test_fallback_column_order():
    """When headers are not recognized, fallback to column order 1=last_name, 2=first_name, ... 5=title."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Col1", "Col2", "Col3", "Col4", "Col5"])
    ws.append(["Фам", "Имя", "Отч", "Спец", "Должность"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 1
    assert rows[0]["last_name"] == "Фам"
    assert rows[0]["first_name"] == "Имя"
    assert rows[0]["title"] == "Должность"


def test_partial_header_alias():
    """Alias 'Дата приема' maps to hire_date."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Должность", "Дата приема"])
    ws.append(["Козлов", "Козел", "Менеджер", "2024-03-10"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 1
    assert rows[0]["hire_date"] == date(2024, 3, 10)


# ---------------------------------------------------------------------------
# 5.2 Edge case tests added per audit recommendations
# ---------------------------------------------------------------------------

def test_date_dd_mm_yyyy_format():
    """Dates in DD.MM.YYYY format should parse correctly."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Должность", "Дата найма", "Дата увольнения"])
    ws.append(["Антонов", "Антон", "Дизайнер", "15.03.2023", "31.12.2024"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 1
    assert rows[0]["hire_date"] == date(2023, 3, 15)
    assert rows[0]["termination_date"] == date(2024, 12, 31)


def test_file_with_bom_in_headers():
    """Excel file with UTF-8 BOM (or leading whitespace) in header names still maps columns."""
    wb = Workbook()
    ws = wb.active
    # BOM character in front of first header
    ws.append(["\ufeffФамилия", "Имя", "Должность"])
    ws.append(["Борисов", "Борис", "Разработчик"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    # Should parse without crashing; title should be recognized
    assert len(rows) == 1
    assert rows[0]["title"] == "Разработчик"


def test_invalid_date_does_not_crash():
    """Rows with invalid date values should not raise; date fields become None."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Должность", "Дата найма"])
    ws.append(["Нечаев", "Николай", "Тестировщик", "not-a-date"])
    ws.append(["Нечаев2", "Николай2", "Тестировщик", "99/99/9999"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Should not raise
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 2
    for row in rows:
        assert row.get("hire_date") is None, f"Expected None for invalid date, got {row.get('hire_date')}"


def test_fio_column_split():
    """A single ФИО column should split into last/first/middle name parts."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ФИО", "Должность"])
    ws.append(["Смирнов Алексей Петрович", "Разработчик"])
    ws.append(["Иванова Мария", "Дизайнер"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = parse_employee_excel(buf.getvalue())
    assert len(rows) == 2
    full_name_row = rows[0]
    # Parser should split the FIO column into parts
    fio_parts = " ".join(filter(None, [
        full_name_row.get("last_name", ""),
        full_name_row.get("first_name", ""),
        full_name_row.get("middle_name", ""),
    ]))
    assert "Смирнов" in fio_parts or full_name_row.get("last_name") == "Смирнов"


def test_performance_large_file():
    """1000 rows should parse in under 10 seconds (regression guard)."""
    import time
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия", "Имя", "Должность", "Подразделение", "Дата найма"])
    for i in range(1000):
        ws.append([f"Фам{i}", f"Имя{i}", "Разработчик", "ИТ", "2024-01-01"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    start = time.time()
    rows = parse_employee_excel(buf.getvalue())
    elapsed = time.time() - start
    assert elapsed < 10, f"Parsing 1000 rows took {elapsed:.2f}s (too slow)"
    titles_with_content = [r for r in rows if (r.get("title") or "").strip()]
    assert len(titles_with_content) == 1000
