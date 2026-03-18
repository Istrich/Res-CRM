"""
Tests for ORM models, Pydantic schema validation, and export service.
"""
import io
from datetime import date

import pytest
from pydantic import ValidationError
from sqlalchemy.exc import IntegrityError

from app.models import Employee, EmployeeProject, SalaryRecord
from app.schemas.assignment import AssignmentCreate
from app.schemas.employee import EmployeeCreate, EmployeeUpdate


# ---------------------------------------------------------------------------
# ORM Model constraints
# ---------------------------------------------------------------------------

class TestModelConstraints:
    def test_employee_termination_before_hire_rejected(self, db):
        emp = Employee(
            title="Dev",
            hire_date=date(2024, 6, 1),
            termination_date=date(2024, 1, 1),
        )
        db.add(emp)
        with pytest.raises(Exception):  # IntegrityError or CheckConstraint
            db.commit()
        db.rollback()

    def test_salary_record_unique_per_month(self, db, make_employee):
        emp = make_employee()
        db.add(SalaryRecord(employee_id=emp.id, year=2024, month=1, salary=100_000))
        db.commit()

        db.add(SalaryRecord(employee_id=emp.id, year=2024, month=1, salary=200_000))
        with pytest.raises(Exception):
            db.commit()
        db.rollback()

    def test_salary_record_month_out_of_range(self, db, make_employee):
        emp = make_employee()
        db.add(SalaryRecord(employee_id=emp.id, year=2024, month=13, salary=100_000))
        with pytest.raises(Exception):
            db.commit()
        db.rollback()

    def test_assignment_rate_must_be_positive(self, db, make_employee, make_project):
        emp = make_employee()
        proj = make_project()
        db.add(EmployeeProject(
            employee_id=emp.id,
            project_id=proj.id,
            rate=0,
            valid_from=date(2024, 1, 1),
        ))
        with pytest.raises(Exception):
            db.commit()
        db.rollback()

    def test_salary_record_cascade_delete(self, db, make_employee, make_salary):
        emp = make_employee()
        make_salary(emp, year=2024, month=1)

        db.delete(emp)
        db.commit()

        records = db.query(SalaryRecord).all()
        assert records == []

    def test_assignment_cascade_delete(self, db, make_employee, make_project, make_assignment):
        emp = make_employee()
        proj = make_project()
        make_assignment(emp, proj)

        db.delete(emp)
        db.commit()

        assignments = db.query(EmployeeProject).all()
        assert assignments == []


class TestEmployeeProperties:
    def test_full_name_all_parts(self):
        emp = Employee(first_name="Иван", last_name="Иванов", middle_name="Петрович", title="Dev")
        assert emp.full_name == "Иванов Иван Петрович"

    def test_full_name_no_middle(self):
        emp = Employee(first_name="Иван", last_name="Иванов", title="Dev")
        assert emp.full_name == "Иванов Иван"

    def test_display_name_employee(self):
        emp = Employee(first_name="Иван", last_name="Иванов", title="Dev", is_position=False)
        assert emp.display_name == "Иванов Иван"

    def test_display_name_position_no_name(self):
        emp = Employee(title="Senior Dev", is_position=True)
        assert "Позиция" in emp.display_name or "Senior Dev" in emp.display_name

    def test_display_name_position_with_name(self):
        emp = Employee(first_name="Иван", last_name="Иванов", title="Dev", is_position=True)
        assert "Иванов" in emp.display_name

    def test_salary_record_total(self):
        rec = SalaryRecord(salary=100_000, kpi_bonus=10_000, fixed_bonus=5_000, one_time_bonus=3_000)
        assert rec.total == 118_000


# ---------------------------------------------------------------------------
# Pydantic schema validation
# ---------------------------------------------------------------------------

class TestSchemaValidation:
    def test_employee_create_valid(self):
        e = EmployeeCreate(title="Dev", first_name="Иван")
        assert e.title == "Dev"

    def test_employee_create_termination_before_hire(self):
        with pytest.raises(ValidationError):
            EmployeeCreate(
                title="Dev",
                hire_date=date(2024, 6, 1),
                termination_date=date(2024, 1, 1),
            )

    def test_employee_create_same_date_ok(self):
        e = EmployeeCreate(
            title="Dev",
            hire_date=date(2024, 1, 1),
            termination_date=date(2024, 1, 1),
        )
        assert e.hire_date == e.termination_date

    def test_assignment_rate_zero_rejected(self):
        with pytest.raises(ValidationError):
            AssignmentCreate(
                employee_id="00000000-0000-0000-0000-000000000001",
                project_id="00000000-0000-0000-0000-000000000002",
                rate=0,
                valid_from=date(2024, 1, 1),
            )

    def test_assignment_rate_above_one_ok(self):
        a = AssignmentCreate(
            employee_id="00000000-0000-0000-0000-000000000001",
            project_id="00000000-0000-0000-0000-000000000002",
            rate=1.5,
            valid_from=date(2024, 1, 1),
        )
        assert a.rate == 1.5


# ---------------------------------------------------------------------------
# Export service
# ---------------------------------------------------------------------------

class TestExportService:
    def test_export_employees_returns_bytes(self, db, full_setup):
        from app.services.export import export_employees
        buf = export_employees(db, 2024)
        assert isinstance(buf, io.BytesIO)
        assert buf.read(4) == b'PK\x03\x04'  # ZIP/xlsx magic bytes

    def test_export_employees_nonempty(self, db, full_setup):
        from app.services.export import export_employees
        buf = export_employees(db, 2024)
        content = buf.read()
        assert len(content) > 1000  # non-trivial xlsx

    def test_export_projects_budget(self, db, full_setup):
        from app.services.export import export_projects_budget
        buf = export_projects_budget(db, 2024)
        assert buf.read(4) == b'PK\x03\x04'

    def test_export_budget_projects(self, db, full_setup):
        from app.services.export import export_budget_projects
        buf = export_budget_projects(db, 2024)
        assert buf.read(4) == b'PK\x03\x04'

    def test_export_payroll_has_two_sheets(self, db, full_setup):
        from app.services.export import export_payroll
        from openpyxl import load_workbook
        buf = export_payroll(db, 2024)
        wb = load_workbook(buf)
        assert len(wb.sheetnames) == 2

    def test_export_empty_db(self, db):
        from app.services.export import export_employees
        buf = export_employees(db, 2024)
        assert buf.read(4) == b'PK\x03\x04'  # still produces valid xlsx


# ---------------------------------------------------------------------------
# Auth service
# ---------------------------------------------------------------------------

class TestAuthService:
    def test_hash_and_verify(self):
        from app.services.auth import hash_password, verify_password
        hashed = hash_password("secret123")
        assert verify_password("secret123", hashed) is True
        assert verify_password("wrong", hashed) is False

    def test_create_and_decode_token(self):
        from app.services.auth import create_access_token, decode_token
        token = create_access_token({"sub": "admin"})
        payload = decode_token(token)
        assert payload is not None
        assert payload["sub"] == "admin"

    def test_bad_token_returns_none(self):
        from app.services.auth import decode_token
        result = decode_token("not.a.real.token")
        assert result is None

    def test_get_or_create_admin(self, db):
        from app.services.auth import get_or_create_admin
        import os
        os.environ.setdefault("ADMIN_USERNAME", "admin")
        os.environ.setdefault("ADMIN_PASSWORD", "admin123")

        user = get_or_create_admin(db)
        assert user.username == "admin"

        # Idempotent
        user2 = get_or_create_admin(db)
        assert user2.id == user.id

    def test_authenticate_user_success(self, db, admin_user):
        from app.services.auth import authenticate_user
        result = authenticate_user(db, "admin", "admin123")
        assert result is not None
        assert result.username == "admin"

    def test_authenticate_user_wrong_password(self, db, admin_user):
        from app.services.auth import authenticate_user
        result = authenticate_user(db, "admin", "wrong")
        assert result is None


# ---------------------------------------------------------------------------
# 5.3 Export cell content validation
# ---------------------------------------------------------------------------

class TestExportCellContent:
    def test_export_employees_headers_and_first_row(self, db, full_setup):
        """export_employees xlsx should have proper headers and data in first row."""
        from app.services.export import export_employees
        from openpyxl import load_workbook
        buf = export_employees(db, 2024)
        wb = load_workbook(buf)
        ws = wb.active

        # Row 1 = headers
        headers = [cell.value for cell in ws[1] if cell.value]
        assert len(headers) > 0, "No headers in export"
        # Should contain employee name or department column
        header_str = " ".join(str(h).lower() for h in headers)
        assert any(keyword in header_str for keyword in ["фамилия", "имя", "сотрудник", "должность", "подразделение"]), \
            f"Expected employee column headers, got: {headers}"

        # Row 2 = first data row
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in data_rows if any(v is not None for v in r)]
        assert len(non_empty) > 0, "Expected at least one data row in employees export"

    def test_export_employees_row_count(self, db, full_setup):
        """Number of data rows matches the number of employees."""
        from app.services.export import export_employees
        from app.models import Employee
        from openpyxl import load_workbook
        buf = export_employees(db, 2024)
        wb = load_workbook(buf)
        ws = wb.active

        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
        emp_count = db.query(Employee).count()
        # Export may include positions too, so at least as many as employees
        assert len(data_rows) >= emp_count

    def test_export_payroll_first_sheet_has_headers(self, db, full_setup):
        """Payroll export first sheet should have headers in row 1."""
        from app.services.export import export_payroll
        from openpyxl import load_workbook
        buf = export_payroll(db, 2024)
        wb = load_workbook(buf)
        ws = wb.worksheets[0]

        headers = [cell.value for cell in ws[1] if cell.value]
        assert len(headers) > 0, "No headers in payroll export"

    def test_export_payroll_employee_name_in_data(self, db, full_setup):
        """Payroll export should contain the employee's name in the data."""
        from app.services.export import export_payroll
        from openpyxl import load_workbook
        buf = export_payroll(db, 2024)
        wb = load_workbook(buf)
        ws = wb.worksheets[0]

        all_values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_values.extend([str(v) for v in row if v is not None])

        emp = full_setup["employee"]
        emp_name_part = emp.last_name or emp.first_name or ""
        assert any(emp_name_part in v for v in all_values), \
            f"Employee name '{emp_name_part}' not found in payroll export"
