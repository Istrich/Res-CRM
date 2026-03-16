"""
Parse Excel file for employee import.
Expected columns (first row): Фамилия, Имя, Отчество, Специализация, Должность,
Подразделение, Дата найма, Дата увольнения, Комментарий.
"""

import io
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Точные и варианты написания заголовков (нижний регистр)
HEADER_MAP = {
    "фамилия": "last_name",
    "surname": "last_name",
    "last name": "last_name",
    "family name": "last_name",
    "имя": "first_name",
    "first name": "first_name",
    "отчество": "middle_name",
    "фио": "fio",
    "специализация": "specialization",
    "должность": "title",
    "подразделение": "department",
    "дата найма": "hire_date",
    "дата увольнения": "termination_date",
    "комментарий": "comment",
}

# Дополнительные варианты (частичное совпадение для распознавания заголовка)
HEADER_ALIASES = [
    ("last_name", ("фамилия", "фам", "surname", "family", "last name", "family name", "lastname")),
    ("first_name", ("имя", "name", "first name", "firstname", "first")),
    ("middle_name", ("отчество", "patronymic", "middle name", "middle")),
    ("fio", ("фио", "ф.и.о", "fio", "полное имя")),
    ("title", ("должность", "position", "title")),
    ("specialization", ("специализация", "specialization", "спец")),
    ("department", ("подразделение", "department", "отдел")),
    ("hire_date", ("дата найма", "hire", "найм", "дата приема")),
    ("termination_date", ("дата увольнения", "увольнение", "termination")),
    ("comment", ("комментарий", "comment", "примечание")),
]


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            try:
                return date(int(s[:4]), int(s[5:7]), int(s[8:10]))
            except (ValueError, TypeError):
                pass
        parts = s.replace(",", ".").split(".")
        if len(parts) == 3:
            try:
                a, b, c = parts[0].strip(), parts[1].strip(), parts[2].strip()
                if len(c) == 4 and len(a) <= 2 and len(b) <= 2:
                    return date(int(c), int(b), int(a))
                if len(a) == 4 and len(b) <= 2 and len(c) <= 2:
                    return date(int(a), int(b), int(c))
            except (ValueError, TypeError):
                pass
    return None


def _normalize_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    # Убрать BOM и лишние пробелы
    if s.startswith("\ufeff"):
        s = s[1:].strip()
    return s


def _match_header_cell(cell_value: str, exact_only: bool = False) -> str | None:
    """
    Возвращает ключ поля по ячейке заголовка.
    exact_only: только точное совпадение с HEADER_MAP (чтобы «Должность» не перебивалась «Код должности»).
    """
    cell_value = _normalize_header(cell_value)
    if not cell_value:
        return None
    if cell_value in HEADER_MAP:
        return HEADER_MAP[cell_value]
    if exact_only:
        return None
    for key, aliases in HEADER_ALIASES:
        for alias in aliases:
            if alias in cell_value or cell_value in alias:
                return key
    return None


def _fill_fio_from_column(row_dict: dict[str, Any]) -> None:
    """
    Если в строке есть значение в колонке ФИО (fio), разбираем его на фамилию, имя, отчество
    и заполняем пустые поля. Формат: «Фамилия Имя Отчество» (пробелы).
    """
    fio_val = row_dict.get("fio")
    if not fio_val or not isinstance(fio_val, str):
        return
    parts = [p.strip() for p in fio_val.strip().split() if p.strip()]
    if not parts:
        return
    if not row_dict.get("last_name") and len(parts) >= 1:
        row_dict["last_name"] = parts[0]
    if not row_dict.get("first_name") and len(parts) >= 2:
        row_dict["first_name"] = parts[1]
    if not row_dict.get("middle_name") and len(parts) >= 3:
        row_dict["middle_name"] = " ".join(parts[2:])


def _cell_value(ws: Any, row: int, col: int) -> Any:
    """
    Возвращает значение ячейки. Для объединённых ячеек openpyxl даёт значение
    только в верхней левой; для остальных возвращает None — подставляем значение
    из верхней ячейки объединённого диапазона.
    """
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return val
    merged = getattr(ws, "merged_cells", None)
    if not merged:
        return None
    coord = f"{get_column_letter(col)}{row}"
    for merged_range in getattr(merged, "ranges", []):
        if coord in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


def parse_employee_excel(file_content: bytes) -> list[dict[str, Any]]:
    """
    Parse first sheet of an Excel file. First row = headers.
    Returns list of dicts with keys: last_name, first_name, middle_name, title,
    department, specialization, hire_date, termination_date, comment.
    """
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    if ws is None:
        return []

    # Сначала точные совпадения (чтобы «Должность» не забирала колонка «Код должности» и т.п.)
    max_col_to_try = max((ws.max_column or 0), 20)
    header_to_col: dict[str, int] = {}
    for c in range(1, max_col_to_try + 1):
        h = ws.cell(row=1, column=c).value
        if h is None and c > (ws.max_column or 0) and not header_to_col:
            break
        key = _match_header_cell(h, exact_only=True)
        if key and key not in header_to_col:
            header_to_col[key] = c
    # Затем частичные (для вариантов вроде «Дата приема», «Отдел»)
    for c in range(1, max_col_to_try + 1):
        h = ws.cell(row=1, column=c).value
        key = _match_header_cell(h, exact_only=False)
        if key and key not in header_to_col:
            header_to_col[key] = c

    if "title" not in header_to_col:
        # Fallback: порядок 1=Фамилия, 2=Имя, 3=Отчество, 4=Специализация, 5=Должность, 6=Подразделение, 7=Дата найма, 8=Дата увольнения, 9=Комментарий
        default_cols = [
            "last_name", "first_name", "middle_name", "specialization", "title",
            "department", "hire_date", "termination_date", "comment",
        ]
        for i, key in enumerate(default_cols, 1):
            if key not in header_to_col:
                header_to_col[key] = i

    if not header_to_col:
        return []

    # Читаем все строки до конца листа; не прерываемся на 2–3 пустых (иначе теряются блоки на «К» и последние строки)
    max_row_to_try = (ws.max_row or 0)
    if max_row_to_try < 2:
        max_row_to_try = 1000
    max_row_to_try = min(max_row_to_try, 5000)
    rows = []
    title_col = header_to_col.get("title", 1)
    empty_title_streak = 0
    STOP_AFTER_CONSECUTIVE_EMPTY = 30  # остановка только после 30 подряд пустых (реальный конец данных)
    for r in range(2, max_row_to_try + 1):
        title_val = _cell_value(ws, r, title_col)
        title_empty = title_val is None or str(title_val).strip() == ""
        if title_empty:
            empty_title_streak += 1
            if empty_title_streak >= STOP_AFTER_CONSECUTIVE_EMPTY and r > 50:
                break
        else:
            empty_title_streak = 0
        row_dict = {}
        for key, col in header_to_col.items():
            val = _cell_value(ws, r, col)
            if key in ("hire_date", "termination_date"):
                row_dict[key] = _parse_date(val)
            else:
                row_dict[key] = (str(val).strip() or None) if val is not None else None
        # Если есть колонка ФИО — разбираем и заполняем фамилию/имя/отчество там, где пусто
        _fill_fio_from_column(row_dict)
        rows.append(row_dict)

    return rows
