"""
Excel export service using openpyxl.
All exports return a BytesIO buffer ready to stream.
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import BudgetProject, Employee, Project, SalaryRecord
from app.services.calc import (
    calc_employee_month_cost,
    employee_active_in_month,
    get_budget_project_summary,
    get_project_budget_summary,
)

MONTH_NAMES = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
               "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")


def _header(ws, row: int, cols: list[str]):
    for i, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ---------------------------------------------------------------------------
# 1. Employees list
# ---------------------------------------------------------------------------

def export_employees(db: Session, year: int) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    month_cols = [f"{m} {year}" for m in MONTH_NAMES]
    headers = [
        "Тип", "ФИО / Позиция", "Должность", "Подразделение", "Специализация",
        "Дата найма", "Дата увольнения", "Проекты", "Ставки",
    ] + month_cols

    _header(ws, 1, headers)

    employees = db.query(Employee).order_by(Employee.last_name).all()

    for r, emp in enumerate(employees, 2):
        fill = ALT_FILL if r % 2 == 0 else None

        projects = [ep.project.name for ep in emp.employee_projects]
        rates = [str(float(ep.rate)) for ep in emp.employee_projects]

        row_data = [
            "Позиция" if emp.is_position else "Сотрудник",
            emp.display_name,
            emp.title,
            emp.department or "",
            emp.specialization or "",
            emp.hire_date.isoformat() if emp.hire_date else "",
            emp.termination_date.isoformat() if emp.termination_date else "",
            ", ".join(projects),
            ", ".join(rates),
        ]

        for month in range(1, 13):
            cost = calc_employee_month_cost(db, emp, year, month)
            row_data.append(cost if cost > 0 else "")

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if fill:
                cell.fill = fill

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 2. Projects budget
# ---------------------------------------------------------------------------

def export_projects_budget(db: Session, year: int) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Бюджеты проектов"

    month_cols = [f"{m} {year}" for m in MONTH_NAMES]
    headers = [
        "Проект", "Бюджетный проект", "Бюджет", "Расход", "Прогноз", "Остаток", "Статус",
    ] + month_cols

    _header(ws, 1, headers)

    projects = db.query(Project).all()

    for r, proj in enumerate(projects, 2):
        summary = get_project_budget_summary(db, proj.id, year)
        fill = ALT_FILL if r % 2 == 0 else None

        from app.models import BudgetSnapshot
        monthly = {
            s.month: float(s.amount)
            for s in db.query(BudgetSnapshot).filter(
                BudgetSnapshot.project_id == proj.id,
                BudgetSnapshot.year == year,
            ).all()
        }

        budget_val = ""
        if proj.budget_project and proj.budget_project.total_budget:
            budget_val = float(proj.budget_project.total_budget)

        row_data = [
            proj.name,
            proj.budget_project.name if proj.budget_project else "",
            budget_val,
            summary["spent"],
            summary["forecast"],
            summary["remaining"] if summary["remaining"] is not None else "",
            {"ok": "В норме", "warning": "Риск", "overrun": "Перерасход"}.get(summary["status"], ""),
        ] + [monthly.get(m, 0) for m in range(1, 13)]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if fill:
                cell.fill = fill

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 3. Budget projects summary
# ---------------------------------------------------------------------------

def export_budget_projects(db: Session, year: int) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Бюджетные проекты"

    headers = [
        "Бюджетный проект", "Год", "Бюджет", "Расход", "Прогноз", "Остаток", "Статус", "Проекты",
    ]
    _header(ws, 1, headers)

    bps = db.query(BudgetProject).filter(BudgetProject.year == year).all()

    for r, bp in enumerate(bps, 2):
        summary = get_budget_project_summary(db, bp.id, year)
        fill = ALT_FILL if r % 2 == 0 else None

        row_data = [
            bp.name,
            bp.year,
            float(bp.total_budget) if bp.total_budget else "",
            summary.get("spent", 0),
            summary.get("forecast", 0),
            summary.get("remaining", ""),
            {"ok": "В норме", "warning": "Риск", "overrun": "Перерасход"}.get(summary.get("status", "ok"), ""),
            ", ".join(p.name for p in bp.projects),
        ]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if fill:
                cell.fill = fill

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 4. Full payroll (ФОТ) — all employees by month
# ---------------------------------------------------------------------------

def export_payroll(db: Session, year: int) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"ФОТ {year}"

    headers = ["ФИО / Позиция", "Должность", "Подразделение"] + [
        f"{m}\nОклад" for m in MONTH_NAMES
    ]

    # Full detail sheet
    ws2 = wb.create_sheet("Детализация")
    detail_headers = [
        "ФИО / Позиция", "Должность", "Подразделение", "Год", "Месяц",
        "Оклад", "KPI", "Фикс. надбавка", "Разовая премия", "Итого",
    ]
    _header(ws2, 1, detail_headers)

    _header(ws, 1, headers)

    employees = db.query(Employee).order_by(Employee.last_name).all()
    detail_row = 2

    for r, emp in enumerate(employees, 2):
        fill = ALT_FILL if r % 2 == 0 else None
        monthly_totals = []

        for month in range(1, 13):
            cost = calc_employee_month_cost(db, emp, year, month)
            monthly_totals.append(cost if cost > 0 else "")

            if cost > 0:
                from app.services.calc import get_salary_for_month
                rec, is_exact = get_salary_for_month(db, emp.id, year, month)
                if rec:
                    one_time = float(rec.one_time_bonus) if is_exact else 0
                    ws2.append([
                        emp.display_name, emp.title, emp.department or "",
                        year, month,
                        float(rec.salary), float(rec.kpi_bonus),
                        float(rec.fixed_bonus), one_time,
                        float(rec.salary) + float(rec.kpi_bonus) + float(rec.fixed_bonus) + one_time,
                    ])
                    detail_row += 1

        row_data = [emp.display_name, emp.title, emp.department or ""] + monthly_totals

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if fill:
                cell.fill = fill

    _autofit(ws)
    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
